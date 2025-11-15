# dispatch_app/management/commands/import_customers.py
import os
from django.core.management.base import BaseCommand
from dispatch_app.models import Customer
import openpyxl

class Command(BaseCommand):
    help = 'Import customers from Excel file'

    def handle(self, *args, **options):
        # 🔸 Update this path to your Excel file
        excel_file_path = r"C:\Users\dispatch\OneDrive - Atyab Food Industries\Documents\Customers.xlsx"

        if not os.path.exists(excel_file_path):
            self.stdout.write(
                self.style.ERROR(f"File not found: {excel_file_path}")
            )
            return

        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            
            # 🔸 Make sure sheet name is exactly "Customers"
            if "Customers" not in workbook.sheetnames:
                self.stdout.write(
                    self.style.ERROR(f'Sheet "Customers" not found. Available sheets: {workbook.sheetnames}')
                )
                return

            worksheet = workbook["Customers"]

            # Assume first row is header
            headers = [cell.value for cell in worksheet[1]]
            self.stdout.write(f"Headers found: {headers}")

            # Map Excel columns to model fields
            # Expected columns: Customer, DispatchTo, Address, Country, ContactNo, ContactPerson, Status
            required_columns = {"Customer"}
            for col in required_columns:
                if col not in headers:
                    self.stdout.write(
                        self.style.ERROR(f"Required column '{col}' missing in Excel.")
                    )
                    return

            created_count = 0
            updated_count = 0
            error_count = 0

            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                # Build dict from row and headers
                row_data = dict(zip(headers, row))

                customer_name = row_data.get("Customer")
                if not customer_name:
                    continue  # Skip rows without Customer name

                # Prepare data
                customer_data = {
                    "DispatchTo": row_data.get("DispatchTo") or "",
                    "Address": row_data.get("Address") or "",
                    "Country": row_data.get("Country") or "",
                    "ContactNo": row_data.get("ContactNo") or "",
                    "ContactPerson": row_data.get("ContactPerson") or "",
                    "Status": row_data.get("Status", True) if row_data.get("Status") is not None else True,
                }

                try:
                    # Update or create
                    obj, created = Customer.objects.update_or_create(
                        Customer=customer_name,
                        defaults=customer_data
                    )

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(f"Row {idx}: Failed to save '{customer_name}' - {e}")
                    )
                    error_count += 1

            self.stdout.write(
                self.style.SUCCESS(
                    f"✅ Import finished!\n"
                    f"Created: {created_count}\n"
                    f"Updated: {updated_count}\n"
                    f"Errors: {error_count}"
                )
            )

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"❌ Fatal error: {e}"))
