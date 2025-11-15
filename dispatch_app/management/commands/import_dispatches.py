# your_app/management/commands/import_dispatches.py
import os
from datetime import datetime
from django.core.management.base import BaseCommand
from django.core.exceptions import ValidationError
from dispatch_app.models import Dispatch, Customer  # ← Replace 'your_app' with your app name
import openpyxl

class Command(BaseCommand):
    help = 'Import dispatches from Excel file'

    def handle(self, *args, **options):
        excel_file_path = r"C:\Users\samad\OneDrive - Atyab Food Industries\Documents\Dispatch.xlsx"

        if not os.path.exists(excel_file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {excel_file_path}"))
            return

        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet_name = "Dispatch"  # Make sure this matches your sheet name
            if sheet_name not in workbook.sheetnames:
                self.stdout.write(self.style.ERROR(f'Sheet "{sheet_name}" not found. Available: {workbook.sheetnames}'))
                return

            worksheet = workbook[sheet_name]
            headers = [cell.value for cell in worksheet[1]]
            self.stdout.write(f"Headers: {headers}")

            # Required fields (must match Excel column names)
            required_cols = {"OrderNo", "Customer", "OrderDate"}
            missing = required_cols - set(headers)
            if missing:
                self.stdout.write(self.style.ERROR(f"Missing required columns: {missing}"))
                return

            created_count = 0
            updated_count = 0
            error_count = 0

            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = dict(zip(headers, row))
                
                order_no = row_data.get("OrderNo")
                customer_name = row_data.get("Customer")
                order_date = row_data.get("OrderDate")

                if not order_no or not customer_name:
                    self.stdout.write(self.style.WARNING(f"Skipping row {idx}: missing OrderNo or Customer"))
                    error_count += 1
                    continue

                # 🔍 Find or skip Customer — tolerant lookup
                def normalize_name(s):
                    if s is None:
                        return None
                    # collapse multiple whitespace into one and strip
                    return ' '.join(str(s).split()).strip()

                norm_name = normalize_name(customer_name)
                customer = None

                # 1) Try case-insensitive exact match with the raw name
                if customer_name:
                    try:
                        customer = Customer.objects.get(Customer__iexact=customer_name)
                    except Customer.DoesNotExist:
                        customer = None

                # 2) Try case-insensitive exact match with normalized name (collapses extra spaces)
                if not customer and norm_name and norm_name != customer_name:
                    try:
                        customer = Customer.objects.get(Customer__iexact=norm_name)
                    except Customer.DoesNotExist:
                        customer = None

                # 3) As a last resort, try a contains lookup (first matching row)
                if not customer and norm_name:
                    qs = Customer.objects.filter(Customer__icontains=norm_name)
                    if qs.exists():
                        customer = qs.first()

                # 4) Python-side normalized comparison fallback
                #    Fetch a small candidate set (by first token) and compare after collapsing whitespace
                if not customer and norm_name:
                    tokens = norm_name.split()
                    if tokens:
                        first = tokens[0]
                        qs = Customer.objects.filter(Customer__icontains=first)
                        for c in qs:
                            if normalize_name(c.Customer).lower() == norm_name.lower():
                                customer = c
                                break

                if not customer:
                    self.stdout.write(
                        self.style.ERROR(f"Row {idx}: Customer '{customer_name}' not found in DB. Skip.")
                    )
                    error_count += 1
                    continue

                # 🗓️ Parse dates safely
                def parse_date(val):
                    if val is None:
                        return None
                    if isinstance(val, datetime):
                        return val.date()
                    if isinstance(val, str):
                        try:
                            return datetime.strptime(val, "%Y-%m-%d").date()
                        except ValueError:
                            try:
                                return datetime.strptime(val, "%d/%m/%Y").date()  # Try DD/MM/YYYY
                            except ValueError:
                                return None
                    return None

                # Prepare dispatch data
                dispatch_data = {
                    "InvoiceNo": row_data.get("InvoiceNo") or None,
                    "Customer": customer,
                    "Address": row_data.get("Address") or None,
                    "Country": row_data.get("Country") or None,
                    "ContactNo": row_data.get("ContactNo") or None,
                    "ContactPerson": row_data.get("ContactPerson") or None,
                    "OrderDate": parse_date(order_date),
                    "LoadingDate": parse_date(row_data.get("LoadingDate")),
                    "DeliveryDate": parse_date(row_data.get("DeliveryDate")),
                    "TransportNo": row_data.get("TransportNo") or None,
                    "DriverName": row_data.get("DriverName") or None,
                    "DriverMobile": row_data.get("DriverMobile") or None,
                    "Seal": row_data.get("Seal") or None,
                    "Status": row_data.get("Status", "draft"),
                }

                # Validate status
                valid_statuses = dict(Dispatch.STATUS_CHOICES).keys()
                if dispatch_data["Status"] not in valid_statuses:
                    dispatch_data["Status"] = "draft"

                # 🔄 Create or update
                try:
                    obj, created = Dispatch.objects.update_or_create(
                        OrderNo=order_no,
                        defaults=dispatch_data
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"Row {idx}: Failed to save - {e}"))
                    error_count += 1

            self.stdout.write(
                self.style.SUCCESS(
                    f"✅ Import finished!\n"
                    f"Created: {created_count}\n"
                    f"Updated: {updated_count}\n"
                    f"Errors: {error_count}"
                )
            )

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"❌ Fatal error: {e}"))