# your_app/management/commands/import_products.py
import os
from django.core.management.base import BaseCommand
from django.conf import settings
from dispatch_app.models import Products  # ← Replace 'your_app' with your app name
import openpyxl

class Command(BaseCommand):
    help = 'Import products from Excel file'

    def handle(self, *args, **options):
        # 🔸 Update this path to your Excel file
        excel_file_path = r"C:\Users\dispatch\OneDrive - Atyab Food Industries\Documents\Products.xlsx"

        if not os.path.exists(excel_file_path):
            self.stdout.write(
                self.style.ERROR(f"File not found: {excel_file_path}")
            )
            return

        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            
            # 🔸 Make sure sheet name is exactly "Products"
            if "Products" not in workbook.sheetnames:
                self.stdout.write(
                    self.style.ERROR('Sheet "Products" not found in Excel file.')
                )
                return

            worksheet = workbook["Products"]

            # Assume first row is header
            headers = [cell.value for cell in worksheet[1]]
            self.stdout.write(f"Headers found: {headers}")

            # Map Excel columns to model fields (adjust if needed)
            # Expected columns: Code, LocalCode, Description, ParPallet, UOM, PacInCtn
            required_columns = {"Code"}
            for col in required_columns:
                if col not in headers:
                    self.stdout.write(
                        self.style.ERROR(f"Required column '{col}' missing in Excel.")
                    )
                    return

            created_count = 0
            updated_count = 0

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                # Build dict from row and headers
                row_data = dict(zip(headers, row))

                code = row_data.get("Code")
                if not code:
                    continue  # Skip rows without Code

                # Convert numeric fields safely
                def to_decimal(val):
                    try:
                        return float(val) if val is not None else None
                    except (ValueError, TypeError):
                        return None

                # Prepare data
                product_data = {
                    "LocalCode": row_data.get("LocalCode") or "",
                    "Description": row_data.get("Description") or "",
                    "UOM": row_data.get("UOM") or "",
                    "ParPallet": to_decimal(row_data.get("ParPallet")),
                    "PacInCtn": to_decimal(row_data.get("PacInCtn")),
                }

                # Update or create
                obj, created = Products.objects.update_or_create(
                    Code=code,
                    defaults=product_data
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            self.stdout.write(
                self.style.SUCCESS(
                    f"Import completed! Created: {created_count}, Updated: {updated_count}"
                )
            )

        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error: {e}"))